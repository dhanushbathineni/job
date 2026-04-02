import json
import sqlite3
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import config
from config import setup_logging

logger = setup_logging("jobbot.db")


SCHEMA_SQL = """
PRAGMA journal_mode=WAL;

CREATE TABLE IF NOT EXISTS jobs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    platform TEXT,
    job_title TEXT,
    company TEXT,
    location TEXT,
    job_description TEXT,
    keywords TEXT,
    application_url TEXT UNIQUE,
    date_posted TEXT,
    date_scraped TEXT,
    salary_range TEXT,
    job_type TEXT,
    status TEXT DEFAULT 'Not Applied',
    match_score INTEGER,
    resume_path TEXT,
    gap_analysis TEXT,
    applied_date TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    updated_at TEXT DEFAULT (datetime('now'))
);

CREATE INDEX IF NOT EXISTS idx_jobs_platform ON jobs(platform);
CREATE INDEX IF NOT EXISTS idx_jobs_status ON jobs(status);
CREATE INDEX IF NOT EXISTS idx_jobs_company ON jobs(company);

CREATE TABLE IF NOT EXISTS scrape_sessions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    platform TEXT NOT NULL,
    search_query TEXT,
    search_location TEXT,
    jobs_found INTEGER DEFAULT 0,
    jobs_new INTEGER DEFAULT 0,
    started_at TEXT,
    finished_at TEXT,
    status TEXT DEFAULT 'running',
    error_message TEXT
);

CREATE TABLE IF NOT EXISTS resume_versions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    job_id INTEGER REFERENCES jobs(id),
    version INTEGER DEFAULT 1,
    file_path TEXT,
    created_at TEXT DEFAULT (datetime('now')),
    claude_model TEXT,
    prompt_tokens INTEGER,
    completion_tokens INTEGER
);

CREATE TABLE IF NOT EXISTS application_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    job_id INTEGER REFERENCES jobs(id),
    action TEXT,
    result TEXT,
    timestamp TEXT DEFAULT (datetime('now')),
    details TEXT
);
"""


_ALLOWED_JOB_COLUMNS = frozenset({
    "platform", "job_title", "company", "location", "job_description",
    "keywords", "application_url", "date_posted", "date_scraped",
    "salary_range", "job_type", "status", "match_score", "resume_path",
    "gap_analysis", "applied_date", "updated_at",
})


class DatabaseManager:
    def __init__(self, db_path: Path = config.DB_PATH):
        self.db_path = str(db_path)

    @contextmanager
    def get_connection(self):
        conn = sqlite3.connect(self.db_path, timeout=30, check_same_thread=False)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def initialize_schema(self):
        with self.get_connection() as conn:
            conn.executescript(SCHEMA_SQL)

    # ── Job operations ────────────────────────────────────────────────────────

    def upsert_job(self, job: dict) -> int:
        """Insert job if application_url is new; return the job id either way."""
        now = datetime.utcnow().isoformat()
        job.setdefault("date_scraped", now)
        if isinstance(job.get("keywords"), list):
            job["keywords"] = json.dumps(job["keywords"])

        cols = [
            "platform", "job_title", "company", "location", "job_description",
            "keywords", "application_url", "date_posted", "date_scraped",
            "salary_range", "job_type",
        ]
        values = [job.get(c) for c in cols]

        with self.get_connection() as conn:
            conn.execute(
                f"INSERT OR IGNORE INTO jobs ({', '.join(cols)}) VALUES ({', '.join(['?']*len(cols))})",
                values,
            )
            row = conn.execute(
                "SELECT id FROM jobs WHERE application_url = ?", (job.get("application_url"),)
            ).fetchone()
            return row["id"] if row else -1

    def get_job(self, job_id: int) -> dict | None:
        with self.get_connection() as conn:
            row = conn.execute("SELECT * FROM jobs WHERE id = ?", (job_id,)).fetchone()
            return dict(row) if row else None

    def get_all_jobs(self, filters: dict | None = None) -> list[dict]:
        sql = "SELECT * FROM jobs"
        params: list[Any] = []
        conditions = []

        if filters:
            if platforms := filters.get("platforms"):
                placeholders = ",".join("?" * len(platforms))
                conditions.append(f"platform IN ({placeholders})")
                params.extend(platforms)
            if statuses := filters.get("statuses"):
                placeholders = ",".join("?" * len(statuses))
                conditions.append(f"status IN ({placeholders})")
                params.extend(statuses)
            if search := filters.get("search"):
                # Escape SQL LIKE wildcards in user-supplied search string
                escaped = search.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
                conditions.append(
                    "(job_title LIKE ? ESCAPE '\\' "
                    "OR company LIKE ? ESCAPE '\\' "
                    "OR keywords LIKE ? ESCAPE '\\')"
                )
                term = f"%{escaped}%"
                params.extend([term, term, term])

        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        sql += " ORDER BY created_at DESC"

        with self.get_connection() as conn:
            rows = conn.execute(sql, params).fetchall()
            return [dict(r) for r in rows]

    def update_job(self, job_id: int, fields: dict):
        # Validate column names against allowlist to prevent SQL injection
        invalid_keys = set(fields.keys()) - _ALLOWED_JOB_COLUMNS - {"updated_at"}
        if invalid_keys:
            raise ValueError(f"Invalid column name(s) for jobs table: {invalid_keys}")
        fields["updated_at"] = datetime.utcnow().isoformat()
        set_clause = ", ".join(f"{k} = ?" for k in fields)
        values = list(fields.values()) + [job_id]
        with self.get_connection() as conn:
            conn.execute(f"UPDATE jobs SET {set_clause} WHERE id = ?", values)

    def update_job_status(self, job_id: int, status: str, applied_date: str | None = None):
        fields: dict = {"status": status}
        if applied_date:
            fields["applied_date"] = applied_date
        self.update_job(job_id, fields)

    def update_job_gap_analysis(self, job_id: int, gap_analysis: dict, match_score: int):
        self.update_job(job_id, {
            "gap_analysis": json.dumps(gap_analysis),
            "match_score": match_score,
        })

    def update_job_resume_path(self, job_id: int, resume_path: str):
        self.update_job(job_id, {"resume_path": resume_path})

    def get_stats(self) -> dict:
        with self.get_connection() as conn:
            total = conn.execute("SELECT COUNT(*) FROM jobs").fetchone()[0]
            by_status = dict(conn.execute(
                "SELECT status, COUNT(*) FROM jobs GROUP BY status"
            ).fetchall())
            by_platform = dict(conn.execute(
                "SELECT platform, COUNT(*) FROM jobs GROUP BY platform"
            ).fetchall())
            recent = conn.execute(
                "SELECT date(date_scraped) as day, COUNT(*) as cnt "
                "FROM jobs GROUP BY day ORDER BY day DESC LIMIT 14"
            ).fetchall()
        return {
            "total": total,
            "by_status": by_status,
            "by_platform": by_platform,
            "recent_by_day": [dict(r) for r in recent],
        }

    # ── Scrape sessions ───────────────────────────────────────────────────────

    def start_scrape_session(self, platform: str, query: str, location: str) -> int:
        with self.get_connection() as conn:
            cur = conn.execute(
                "INSERT INTO scrape_sessions (platform, search_query, search_location, started_at) "
                "VALUES (?, ?, ?, ?)",
                (platform, query, location, datetime.utcnow().isoformat()),
            )
            return cur.lastrowid

    def finish_scrape_session(self, session_id: int, jobs_found: int, jobs_new: int):
        with self.get_connection() as conn:
            conn.execute(
                "UPDATE scrape_sessions SET status='completed', finished_at=?, jobs_found=?, jobs_new=? "
                "WHERE id=?",
                (datetime.utcnow().isoformat(), jobs_found, jobs_new, session_id),
            )

    def fail_scrape_session(self, session_id: int, error: str):
        with self.get_connection() as conn:
            conn.execute(
                "UPDATE scrape_sessions SET status='failed', finished_at=?, error_message=? WHERE id=?",
                (datetime.utcnow().isoformat(), error, session_id),
            )

    # ── Resume versions ───────────────────────────────────────────────────────

    def save_resume_version(self, job_id: int, file_path: str, tokens: dict) -> int:
        with self.get_connection() as conn:
            version = (conn.execute(
                "SELECT COUNT(*) FROM resume_versions WHERE job_id=?", (job_id,)
            ).fetchone()[0] or 0) + 1
            cur = conn.execute(
                "INSERT INTO resume_versions (job_id, version, file_path, claude_model, prompt_tokens, completion_tokens) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (job_id, version, file_path, config.OLLAMA_MODEL,
                 tokens.get("input_tokens", 0), tokens.get("output_tokens", 0)),
            )
            return cur.lastrowid

    # ── Application log ───────────────────────────────────────────────────────

    def log_action(self, job_id: int, action: str, result: str, details: dict | None = None):
        with self.get_connection() as conn:
            conn.execute(
                "INSERT INTO application_log (job_id, action, result, details) VALUES (?, ?, ?, ?)",
                (job_id, action, result, json.dumps(details) if details else None),
            )

    # ── Excel export ──────────────────────────────────────────────────────────

    def export_to_excel(self, output_path: Path = config.EXCEL_PATH):
        jobs = self.get_all_jobs()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Jobs"

        headers = [
            "ID", "Platform", "Title", "Company", "Location", "Status",
            "Match Score", "Keywords", "Salary", "Job Type", "Date Posted",
            "Date Scraped", "Application URL", "Resume Path", "Applied Date",
        ]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        status_fills = {
            "Not Applied": PatternFill("solid", fgColor="FFE5E5"),
            "In Progress": PatternFill("solid", fgColor="FFFDE7"),
            "Applied": PatternFill("solid", fgColor="E8F5E9"),
        }

        for row_idx, job in enumerate(jobs, 2):
            keywords = job.get("keywords") or "[]"
            if isinstance(keywords, str):
                try:
                    keywords = ", ".join(json.loads(keywords))
                except (json.JSONDecodeError, TypeError) as e:
                    logger.warning("Failed to parse keywords JSON for job %s: %s", job.get("id"), e)

            row_data = [
                job["id"], job.get("platform"), job.get("job_title"), job.get("company"),
                job.get("location"), job.get("status"), job.get("match_score"),
                keywords, job.get("salary_range"), job.get("job_type"),
                job.get("date_posted"), job.get("date_scraped"),
                job.get("application_url"), job.get("resume_path"), job.get("applied_date"),
            ]
            fill = status_fills.get(job.get("status", ""), None)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if fill:
                    cell.fill = fill

        # Auto-fit columns (approximate)
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        wb.save(output_path)
        return output_path
